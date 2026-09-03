import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_DIR = SCRIPT_DIR / "inputs" / "Ablation" / "demo"
DEFAULT_OUTPUT = DEFAULT_INPUT_DIR / "agent_collision_rate_table.xlsx"


def _seed_sort_key(path: Path):
    match = re.search(r"seed_?(\d+)", path.parent.name)
    return int(match.group(1)) if match else path.parent.name


def _load_collision_series(path: Path):
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if "collision_agents_rate_list" in data:
        key = "collision_agents_rate_list"
        scale = 1.0
    elif "agent_collision_rate_list" in data:
        key = "agent_collision_rate_list"
        scale = 0.01
    else:
        raise KeyError(
            f"{path} must contain collision_agents_rate_list or agent_collision_rate_list"
        )

    values = [float(value) * scale for value in data[key]]
    return {
        "seed": path.parent.name,
        "file": path.name,
        "key": key,
        "scale_to_ratio": scale,
        "values": values,
    }


def _format_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(max_len + 2, 12), 36
        )


def main():
    parser = argparse.ArgumentParser(
        description="Export ablation collision rate series to an Excel table."
    )
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    json_paths = sorted(args.input_dir.glob("seed*/reward*_data.json"), key=_seed_sort_key)
    if not json_paths:
        raise FileNotFoundError(f"No seed*/reward*_data.json found in {args.input_dir}")

    runs = [_load_collision_series(path) for path in json_paths]
    max_len = max(len(run["values"]) for run in runs)

    workbook = Workbook()
    series_sheet = workbook.active
    series_sheet.title = "collision_series_ratio"
    series_sheet.append(["episode"] + [run["seed"] for run in runs])

    for idx in range(max_len):
        row = [idx + 1]
        for run in runs:
            row.append(run["values"][idx] if idx < len(run["values"]) else None)
        series_sheet.append(row)

    for row in series_sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "0.000000"
    _format_sheet(series_sheet)

    summary_sheet = workbook.create_sheet("summary")
    summary_sheet.append(
        [
            "seed",
            "file",
            "source_key",
            "scale_to_ratio",
            "count",
            "mean_ratio",
            "std_ratio",
            "min_ratio",
            "max_ratio",
        ]
    )
    for run in runs:
        values = run["values"]
        mean = sum(values) / len(values)
        variance = sum((value - mean) ** 2 for value in values) / len(values)
        summary_sheet.append(
            [
                run["seed"],
                run["file"],
                run["key"],
                run["scale_to_ratio"],
                len(values),
                mean,
                variance**0.5,
                min(values),
                max(values),
            ]
        )

    for row in summary_sheet.iter_rows(min_row=2, min_col=4):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000000"
    _format_sheet(summary_sheet)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"[DONE] Saved {args.output}")


if __name__ == "__main__":
    main()
