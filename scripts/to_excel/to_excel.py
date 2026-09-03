import pandas as pd
import re
import os
import numpy as np
import matplotlib.pyplot as plt

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "outputs")

# Clover、Weave、Bypass、Merge

# ==========================================
# 1. 原始日志数据
# ==========================================
# log_content = """

# MFPO:
# Scenario: Clover
# [LOG] Agent-agent collision rate [%]: mean=1.71, min=0.42, max=3.00
# [LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
# [LOG] Total collision rate [%]: mean=1.71, min=0.42, max=3.00
# [LOG] Relative centerline deviation [%]: mean=39.60, min=38.23, max=41.83
# [LOG] Relative average speed [%]: mean=89.01, min=89.35, max=90.75
# [LOG] Smoothness [%]: mean=94.79, min=94.53, max=94.99
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Weave
# [LOG] Agent-agent collision rate [%]: mean=5.46, min=4.09, max=7.09
# [LOG] Agent-lanelet collision rate [%]: mean=2.29, min=1.50, max=3.00
# [LOG] Total collision rate [%]: mean=7.75, min=6.17, max=9.26
# [LOG] Relative centerline deviation [%]: mean=34.90, min=33.92, max=35.86
# [LOG] Relative average speed [%]: mean=84.99, min=87.65, max=88.26
# [LOG] Smoothness [%]: mean=94.79, min=94.64, max=95.00
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Bypass
# [LOG] Agent-agent collision rate [%]: mean=4.37, min=3.34, max=5.67
# [LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
# [LOG] Total collision rate [%]: mean=4.37, min=3.34, max=5.67
# [LOG] Relative centerline deviation [%]: mean=34.96, min=34.52, max=35.65
# [LOG] Relative average speed [%]: mean=86.37, min=87.24, max=87.52
# [LOG] Smoothness [%]: mean=96.43, min=96.34, max=96.53
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Merge
# [LOG] Agent-agent collision rate [%]: mean=4.80, min=3.84, max=5.59
# [LOG] Agent-lanelet collision rate [%]: mean=0.78, min=0.33, max=1.33
# [LOG] Total collision rate [%]: mean=5.57, min=4.59, max=6.26
# [LOG] Relative centerline deviation [%]: mean=37.74, min=36.65, max=38.70
# [LOG] Relative average speed [%]: mean=83.55, min=87.27, max=87.79
# [LOG] Smoothness [%]: mean=94.85, min=94.71, max=95.07
# [LOG] Composite scores: -1.00
# =========================================

# Sigma:
# Scenario: Clover
# [LOG] Agent-agent collision rate [%]: mean=1.67, min=0.83, max=2.25
# [LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
# [LOG] Total collision rate [%]: mean=1.67, min=0.83, max=2.25
# [LOG] Relative centerline deviation [%]: mean=38.41, min=37.00, max=40.22
# [LOG] Relative average speed [%]: mean=88.39, min=89.18, max=92.26
# [LOG] Smoothness [%]: mean=95.17, min=94.82, max=95.50
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Weave
# [LOG] Agent-agent collision rate [%]: mean=5.45, min=4.42, max=6.51
# [LOG] Agent-lanelet collision rate [%]: mean=0.95, min=0.50, max=1.67
# [LOG] Total collision rate [%]: mean=6.40, min=5.25, max=7.84
# [LOG] Relative centerline deviation [%]: mean=28.96, min=28.11, max=30.55
# [LOG] Relative average speed [%]: mean=86.00, min=91.84, max=92.24
# [LOG] Smoothness [%]: mean=95.15, min=94.88, max=95.32
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Bypass
# [LOG] Agent-agent collision rate [%]: mean=4.28, min=3.25, max=5.59
# [LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
# [LOG] Total collision rate [%]: mean=4.28, min=3.25, max=5.59
# [LOG] Relative centerline deviation [%]: mean=25.59, min=25.12, max=26.19
# [LOG] Relative average speed [%]: mean=87.68, min=91.59, max=91.81
# [LOG] Smoothness [%]: mean=97.39, min=97.32, max=97.45
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Merge
# [LOG] Agent-agent collision rate [%]: mean=5.02, min=4.17, max=5.92
# [LOG] Agent-lanelet collision rate [%]: mean=0.03, min=0.00, max=0.25
# [LOG] Total collision rate [%]: mean=5.06, min=4.25, max=5.92
# [LOG] Relative centerline deviation [%]: mean=31.64, min=30.87, max=32.43
# [LOG] Relative average speed [%]: mean=85.54, min=91.36, max=91.81
# [LOG] Smoothness [%]: mean=94.79, min=94.65, max=94.98
# [LOG] Composite scores: -1.00
# =========================================


# XP:
# Scenario: Clover
# [LOG] Agent-agent collision rate [%]: mean=0.19, min=0.00, max=0.42
# [LOG] Agent-lanelet collision rate [%]: mean=0.03, min=0.00, max=0.08
# [LOG] Total collision rate [%]: mean=0.22, min=0.00, max=0.50
# [LOG] Relative centerline deviation [%]: mean=50.71, min=43.19, max=59.64
# [LOG] Relative average speed [%]: mean=85.25, min=82.75, max=88.45
# [LOG] Smoothness [%]: mean=94.61, min=94.00, max=95.35
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Weave
# [LOG] Agent-agent collision rate [%]: mean=0.69, min=0.25, max=1.17
# [LOG] Agent-lanelet collision rate [%]: mean=0.98, min=0.50, max=1.33
# [LOG] Total collision rate [%]: mean=1.66, min=1.00, max=2.25
# [LOG] Relative centerline deviation [%]: mean=25.58, min=24.74, max=26.50
# [LOG] Relative average speed [%]: mean=81.49, min=79.92, max=83.01
# [LOG] Smoothness [%]: mean=94.12, min=93.74, max=94.45
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Bypass
# [LOG] Agent-agent collision rate [%]: mean=0.24, min=0.00, max=0.58
# [LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
# [LOG] Total collision rate [%]: mean=0.24, min=0.00, max=0.58
# [LOG] Relative centerline deviation [%]: mean=28.47, min=27.91, max=29.11
# [LOG] Relative average speed [%]: mean=81.90, min=79.28, max=84.46
# [LOG] Smoothness [%]: mean=94.80, min=94.37, max=95.20
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Merge
# [LOG] Agent-agent collision rate [%]: mean=0.56, min=0.08, max=1.25
# [LOG] Agent-lanelet collision rate [%]: mean=0.07, min=0.00, max=0.17
# [LOG] Total collision rate [%]: mean=0.62, min=0.08, max=1.25
# [LOG] Relative centerline deviation [%]: mean=30.91, min=30.23, max=31.86
# [LOG] Relative average speed [%]: mean=79.24, min=77.46, max=81.93
# [LOG] Smoothness [%]: mean=93.14, min=92.58, max=93.54
# [LOG] Composite scores: -1.00
# =========================================


# TSC:
# Scenario: Clover
# [LOG] Agent-agent collision rate [%]: mean=0.04, min=0.00, max=0.17
# [LOG] Agent-lanelet collision rate [%]: mean=0.03, min=0.00, max=0.17
# [LOG] Total collision rate [%]: mean=0.07, min=0.00, max=0.33
# [LOG] Relative centerline deviation [%]: mean=48.50, min=44.68, max=51.99
# [LOG] Relative average speed [%]: mean=88.31, min=85.94, max=89.74
# [LOG] Smoothness [%]: mean=94.66, min=93.64, max=95.26
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Weave
# [LOG] Agent-agent collision rate [%]: mean=0.18, min=0.00, max=0.50
# [LOG] Agent-lanelet collision rate [%]: mean=0.81, min=0.33, max=1.33
# [LOG] Total collision rate [%]: mean=0.99, min=0.42, max=1.67
# [LOG] Relative centerline deviation [%]: mean=36.56, min=35.53, max=37.72
# [LOG] Relative average speed [%]: mean=83.81, min=81.10, max=85.08
# [LOG] Smoothness [%]: mean=93.66, min=93.37, max=93.89
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Bypass
# [LOG] Agent-agent collision rate [%]: mean=0.03, min=0.00, max=0.17
# [LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
# [LOG] Total collision rate [%]: mean=0.03, min=0.00, max=0.17
# [LOG] Relative centerline deviation [%]: mean=34.64, min=34.20, max=35.01
# [LOG] Relative average speed [%]: mean=84.78, min=84.00, max=85.75
# [LOG] Smoothness [%]: mean=94.20, min=93.72, max=94.57
# [LOG] Composite scores: -1.00
# =========================================
# Scenario: Merge
# [LOG] Agent-agent collision rate [%]: mean=0.08, min=0.00, max=0.25
# [LOG] Agent-lanelet collision rate [%]: mean=0.03, min=0.00, max=0.08
# [LOG] Total collision rate [%]: mean=0.10, min=0.00, max=0.33
# [LOG] Relative centerline deviation [%]: mean=37.28, min=36.28, max=39.15
# [LOG] Relative average speed [%]: mean=82.58, min=77.78, max=81.02
# [LOG] Smoothness [%]: mean=91.80, min=90.89, max=92.54
# [LOG] Composite scores: -1.00
# =========================================

# """

log_content = """

MFPO:
Scenario: Clover
[LOG] Agent-agent collision rate [%]: mean=1.71, min=0.42, max=3.00
[LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
[LOG] Total collision rate [%]: mean=1.71, min=0.42, max=3.00
[LOG] Relative centerline deviation [%]: mean=39.60, min=38.23, max=41.83
[LOG] Relative average speed [%]: mean=90.01, min=89.35, max=90.75
[LOG] Relative average speed with collision penalty [%]: mean=86.59
[LOG] Smoothness [%]: mean=5.21, min=5.01, max=5.47
[LOG] Smoothness with collision penalty [%]: mean=6.92
[LOG] Smoothness longitudinal [%]: mean=0.92
[LOG] Smoothness lateral [%]: mean=9.49
[LOG] Composite scores: -1.00
=========================================
Scenario: Weave
[LOG] Agent-agent collision rate [%]: mean=5.46, min=4.09, max=7.09
[LOG] Agent-lanelet collision rate [%]: mean=2.29, min=1.50, max=3.00
[LOG] Total collision rate [%]: mean=7.75, min=6.17, max=9.26
[LOG] Relative centerline deviation [%]: mean=34.90, min=33.92, max=35.86
[LOG] Relative average speed [%]: mean=87.99, min=87.65, max=88.26
[LOG] Relative average speed with collision penalty [%]: mean=72.48
[LOG] Smoothness [%]: mean=5.21, min=5.00, max=5.36
[LOG] Smoothness with collision penalty [%]: mean=12.96
[LOG] Smoothness longitudinal [%]: mean=1.69
[LOG] Smoothness lateral [%]: mean=8.72
[LOG] Composite scores: -1.00
=========================================
Scenario: Merge
[LOG] Agent-agent collision rate [%]: mean=4.37, min=3.34, max=5.67
[LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
[LOG] Total collision rate [%]: mean=4.37, min=3.34, max=5.67
[LOG] Relative centerline deviation [%]: mean=34.96, min=34.52, max=35.65
[LOG] Relative average speed [%]: mean=87.37, min=87.24, max=87.52
[LOG] Relative average speed with collision penalty [%]: mean=78.62
[LOG] Smoothness [%]: mean=3.57, min=3.47, max=3.66
[LOG] Smoothness with collision penalty [%]: mean=7.95
[LOG] Smoothness longitudinal [%]: mean=1.47
[LOG] Smoothness lateral [%]: mean=5.67
[LOG] Composite scores: -1.00
=========================================
Scenario: Bypass
[LOG] Agent-agent collision rate [%]: mean=4.80, min=3.84, max=5.59
[LOG] Agent-lanelet collision rate [%]: mean=0.78, min=0.33, max=1.33
[LOG] Total collision rate [%]: mean=5.57, min=4.59, max=6.26
[LOG] Relative centerline deviation [%]: mean=37.74, min=36.65, max=38.70
[LOG] Relative average speed [%]: mean=87.55, min=87.27, max=87.79
[LOG] Relative average speed with collision penalty [%]: mean=76.41
[LOG] Smoothness [%]: mean=5.15, min=4.93, max=5.29
[LOG] Smoothness with collision penalty [%]: mean=10.73
[LOG] Smoothness longitudinal [%]: mean=1.36
[LOG] Smoothness lateral [%]: mean=8.95
[LOG] Composite scores: -1.00
=========================================



SigmaRL:
Scenario: Clover
[LOG] Agent-agent collision rate [%]: mean=1.67, min=0.83, max=2.25
[LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
[LOG] Total collision rate [%]: mean=1.67, min=0.83, max=2.25
[LOG] Relative centerline deviation [%]: mean=38.41, min=37.00, max=40.22
[LOG] Relative average speed [%]: mean=90.39, min=89.18, max=92.26
[LOG] Relative average speed with collision penalty [%]: mean=87.06
[LOG] Smoothness [%]: mean=4.83, min=4.50, max=5.18
[LOG] Smoothness with collision penalty [%]: mean=6.50
[LOG] Smoothness longitudinal [%]: mean=0.91
[LOG] Smoothness lateral [%]: mean=8.75
[LOG] Composite scores: -1.00
=========================================
Scenario: Weave
[LOG] Agent-agent collision rate [%]: mean=5.45, min=4.42, max=6.51
[LOG] Agent-lanelet collision rate [%]: mean=0.95, min=0.50, max=1.67
[LOG] Total collision rate [%]: mean=6.40, min=5.25, max=7.84
[LOG] Relative centerline deviation [%]: mean=28.96, min=28.11, max=30.55
[LOG] Relative average speed [%]: mean=92.00, min=91.84, max=92.24
[LOG] Relative average speed with collision penalty [%]: mean=79.19
[LOG] Smoothness [%]: mean=4.85, min=4.68, max=5.12
[LOG] Smoothness with collision penalty [%]: mean=11.25
[LOG] Smoothness longitudinal [%]: mean=1.19
[LOG] Smoothness lateral [%]: mean=8.51
[LOG] Composite scores: -1.00
=========================================
Scenario: Merge
[LOG] Agent-agent collision rate [%]: mean=4.28, min=3.25, max=5.59
[LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
[LOG] Total collision rate [%]: mean=4.28, min=3.25, max=5.59
[LOG] Relative centerline deviation [%]: mean=25.59, min=25.12, max=26.19
[LOG] Relative average speed [%]: mean=91.68, min=91.59, max=91.81
[LOG] Relative average speed with collision penalty [%]: mean=83.11
[LOG] Smoothness [%]: mean=2.61, min=2.55, max=2.68
[LOG] Smoothness with collision penalty [%]: mean=6.90
[LOG] Smoothness longitudinal [%]: mean=0.95
[LOG] Smoothness lateral [%]: mean=4.28
[LOG] Composite scores: -1.00
=========================================
Scenario: Bypass
[LOG] Agent-agent collision rate [%]: mean=5.02, min=4.17, max=5.92
[LOG] Agent-lanelet collision rate [%]: mean=0.03, min=0.00, max=0.25
[LOG] Total collision rate [%]: mean=5.06, min=4.25, max=5.92
[LOG] Relative centerline deviation [%]: mean=31.64, min=30.87, max=32.43
[LOG] Relative average speed [%]: mean=91.54, min=91.36, max=91.81
[LOG] Relative average speed with collision penalty [%]: mean=78.43
[LOG] Smoothness [%]: mean=5.21, min=5.02, max=5.35
[LOG] Smoothness with collision penalty [%]: mean=10.27
[LOG] Smoothness longitudinal [%]: mean=1.12
[LOG] Smoothness lateral [%]: mean=9.31
[LOG] Composite scores: -1.00
=========================================



XP-MARL:
Scenario: Clover
[LOG] Agent-agent collision rate [%]: mean=0.19, min=0.00, max=0.42
[LOG] Agent-lanelet collision rate [%]: mean=0.03, min=0.00, max=0.08
[LOG] Total collision rate [%]: mean=0.22, min=0.00, max=0.50
[LOG] Relative centerline deviation [%]: mean=50.71, min=43.19, max=59.64
[LOG] Relative average speed [%]: mean=85.25, min=82.75, max=88.45
[LOG] Relative average speed with collision penalty [%]: mean=84.80
[LOG] Smoothness [%]: mean=5.39, min=4.65, max=6.00
[LOG] Smoothness with collision penalty [%]: mean=5.61
[LOG] Smoothness longitudinal [%]: mean=2.05
[LOG] Smoothness lateral [%]: mean=8.73
[LOG] Composite scores: -1.00
=========================================
Scenario: Weave
[LOG] Agent-agent collision rate [%]: mean=0.69, min=0.25, max=1.17
[LOG] Agent-lanelet collision rate [%]: mean=0.98, min=0.50, max=1.33
[LOG] Total collision rate [%]: mean=1.66, min=1.00, max=2.25
[LOG] Relative centerline deviation [%]: mean=25.58, min=24.74, max=26.50
[LOG] Relative average speed [%]: mean=81.49, min=79.92, max=83.01
[LOG] Relative average speed with collision penalty [%]: mean=78.16
[LOG] Smoothness [%]: mean=5.88, min=5.55, max=6.26
[LOG] Smoothness with collision penalty [%]: mean=7.54
[LOG] Smoothness longitudinal [%]: mean=3.92
[LOG] Smoothness lateral [%]: mean=7.84
[LOG] Composite scores: -1.00
=========================================
Scenario: Merge
[LOG] Agent-agent collision rate [%]: mean=0.24, min=0.00, max=0.58
[LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
[LOG] Total collision rate [%]: mean=0.24, min=0.00, max=0.58
[LOG] Relative centerline deviation [%]: mean=28.47, min=27.91, max=29.11
[LOG] Relative average speed [%]: mean=81.90, min=79.28, max=84.46
[LOG] Relative average speed with collision penalty [%]: mean=81.41
[LOG] Smoothness [%]: mean=5.20, min=4.80, max=5.63
[LOG] Smoothness with collision penalty [%]: mean=5.45
[LOG] Smoothness longitudinal [%]: mean=5.33
[LOG] Smoothness lateral [%]: mean=5.08
[LOG] Composite scores: -1.00
=========================================
Scenario: Bypass
[LOG] Agent-agent collision rate [%]: mean=0.56, min=0.08, max=1.25
[LOG] Agent-lanelet collision rate [%]: mean=0.07, min=0.00, max=0.17
[LOG] Total collision rate [%]: mean=0.62, min=0.08, max=1.25
[LOG] Relative centerline deviation [%]: mean=30.91, min=30.23, max=31.86
[LOG] Relative average speed [%]: mean=79.24, min=77.46, max=81.93
[LOG] Relative average speed with collision penalty [%]: mean=77.99
[LOG] Smoothness [%]: mean=6.86, min=6.46, max=7.42
[LOG] Smoothness with collision penalty [%]: mean=7.48
[LOG] Smoothness longitudinal [%]: mean=5.32
[LOG] Smoothness lateral [%]: mean=8.39
[LOG] Composite scores: -1.00
=========================================




TSC:
Scenario: Clover
[LOG] Agent-agent collision rate [%]: mean=0.04, min=0.00, max=0.17
[LOG] Agent-lanelet collision rate [%]: mean=0.03, min=0.00, max=0.17
[LOG] Total collision rate [%]: mean=0.07, min=0.00, max=0.33
[LOG] Relative centerline deviation [%]: mean=48.50, min=44.68, max=51.99
[LOG] Relative average speed [%]: mean=88.31, min=85.94, max=89.74
[LOG] Relative average speed with collision penalty [%]: mean=88.17
[LOG] Smoothness [%]: mean=5.34, min=4.74, max=6.36
[LOG] Smoothness with collision penalty [%]: mean=5.42
[LOG] Smoothness longitudinal [%]: mean=1.87
[LOG] Smoothness lateral [%]: mean=8.81
[LOG] Composite scores: -1.00
=========================================
Scenario: Weave
[LOG] Agent-agent collision rate [%]: mean=0.18, min=0.00, max=0.50
[LOG] Agent-lanelet collision rate [%]: mean=0.81, min=0.33, max=1.33
[LOG] Total collision rate [%]: mean=0.99, min=0.42, max=1.67
[LOG] Relative centerline deviation [%]: mean=36.56, min=35.53, max=37.72
[LOG] Relative average speed [%]: mean=83.81, min=81.10, max=85.08
[LOG] Relative average speed with collision penalty [%]: mean=81.83
[LOG] Smoothness [%]: mean=6.34, min=6.11, max=6.63
[LOG] Smoothness with collision penalty [%]: mean=7.33
[LOG] Smoothness longitudinal [%]: mean=4.30
[LOG] Smoothness lateral [%]: mean=8.39
[LOG] Composite scores: -1.00
=========================================
Scenario: Merge
[LOG] Agent-agent collision rate [%]: mean=0.03, min=0.00, max=0.17
[LOG] Agent-lanelet collision rate [%]: mean=0.00, min=0.00, max=0.00
[LOG] Total collision rate [%]: mean=0.03, min=0.00, max=0.17
[LOG] Relative centerline deviation [%]: mean=34.64, min=34.20, max=35.01
[LOG] Relative average speed [%]: mean=84.78, min=84.00, max=85.75
[LOG] Relative average speed with collision penalty [%]: mean=84.71
[LOG] Smoothness [%]: mean=5.80, min=5.43, max=6.28
[LOG] Smoothness with collision penalty [%]: mean=5.84
[LOG] Smoothness longitudinal [%]: mean=5.06
[LOG] Smoothness lateral [%]: mean=6.55
[LOG] Composite scores: -1.00
=========================================
Scenario: Bypass
[LOG] Agent-agent collision rate [%]: mean=0.08, min=0.00, max=0.25
[LOG] Agent-lanelet collision rate [%]: mean=0.03, min=0.00, max=0.08
[LOG] Total collision rate [%]: mean=0.10, min=0.00, max=0.33
[LOG] Relative centerline deviation [%]: mean=37.28, min=36.28, max=39.15
[LOG] Relative average speed [%]: mean=79.58, min=77.78, max=81.02
[LOG] Relative average speed with collision penalty [%]: mean=79.38
[LOG] Smoothness [%]: mean=8.20, min=7.46, max=9.11
[LOG] Smoothness with collision penalty [%]: mean=8.30
[LOG] Smoothness longitudinal [%]: mean=7.34
[LOG] Smoothness lateral [%]: mean=9.06
[LOG] Composite scores: -1.00
=========================================




"""


def generate_excel_final_norm(log_text, output_file=None):
    if output_file is None:
        output_file = os.path.join(OUTPUT_DIR, "Comparison_Result_Optimized.xlsx")
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    data = []
    current_algo = None
    current_scenario = None

    # 1. 定义需要反转（值越小越好 -> 分数越高越好）的指标关键词
    # 包含这些关键词的指标，归一化公式为: 1 - (x - min)/(max - min)
    negative_keywords = ["collision", "deviation", "smoothness"]

    # 解析数据
    lines = log_text.strip().split("\n")
    for line in lines:
        line = line.strip()
        if not line:
            continue

        if (
            (line.endswith(":") or line.endswith("："))
            and "Scenario" not in line
            and "[LOG]" not in line
        ):
            current_algo = line.strip(" :：")
            if "XP-MARL" in current_algo:
                current_algo = "XP-MARL"
            continue

        if line.startswith("Scenario:"):
            current_scenario = line.split(":", 1)[1].strip()
            continue

        if (
            line.startswith("[LOG]")
            and "Relative average speed with collision penalty" in line
        ):
            metric_name = "Relative average speed"
            mean_m = re.search(r"mean=([\d\.]+)", line)
            if mean_m:
                common_info = {
                    "Scenario": current_scenario,
                    "Metric": metric_name,
                    "Algorithm": current_algo,
                }
                data.append(
                    {
                        **common_info,
                        "Stat": "Mean",
                        "Value": float(mean_m.group(1)),
                    }
                )
            continue

        if line.startswith("[LOG]") and "Smoothness with collision penalty" in line:
            metric_name = "Smoothness"
            mean_m = re.search(r"mean=([\d\.]+)", line)
            if mean_m:
                common_info = {
                    "Scenario": current_scenario,
                    "Metric": metric_name,
                    "Algorithm": current_algo,
                }
                data.append(
                    {
                        **common_info,
                        "Stat": "Mean",
                        "Value": float(mean_m.group(1)),
                    }
                )
            continue

        if line.startswith("[LOG]") and "mean=" in line:
            if (
                "Relative average speed [%]" in line
                and "with collision penalty" not in line
            ):
                continue
            if "Smoothness [%]" in line and "with collision penalty" not in line:
                continue

            metric_part, values_part = line.split(":", 1)
            metric_name = metric_part.replace("[LOG]", "").replace("[%]", "").strip()

            mean_m = re.search(r"mean=([\d\.]+)", values_part)
            min_m = re.search(r"min=([\d\.]+)", values_part)
            max_m = re.search(r"max=([\d\.]+)", values_part)

            if mean_m and min_m and max_m:
                common_info = {
                    "Scenario": current_scenario,
                    "Metric": metric_name,
                    "Algorithm": current_algo,
                }
                data.append(
                    {
                        **common_info,
                        "Stat": "Mean",
                        "Value": float(mean_m.group(1)),
                    }
                )
                data.append(
                    {
                        **common_info,
                        "Stat": "Min",
                        "Value": float(min_m.group(1)),
                    }
                )
                data.append(
                    {
                        **common_info,
                        "Stat": "Max",
                        "Value": float(max_m.group(1)),
                    }
                )

    df = pd.DataFrame(data)

    # ---------------------------------------------------------
    # 优化后的归一化逻辑
    # ---------------------------------------------------------

    # 1. 获取全局最小值和最大值 (基于所有算法的Min/Max)
    min_stats = (
        df[df["Stat"] == "Mean"]
        .groupby(["Scenario", "Metric"])["Value"]
        .min()
        .reset_index()
    )
    min_stats.rename(columns={"Value": "Global_Min"}, inplace=True)

    max_stats = (
        df[df["Stat"] == "Mean"]
        .groupby(["Scenario", "Metric"])["Value"]
        .max()
        .reset_index()
    )
    max_stats.rename(columns={"Value": "Global_Max"}, inplace=True)

    global_bounds = pd.merge(min_stats, max_stats, on=["Scenario", "Metric"])

    # 2. 合并统计值
    mean_df = df[df["Stat"] == "Mean"].copy()
    merged_df = pd.merge(mean_df, global_bounds, on=["Scenario", "Metric"], how="left")

    # 3. 计算归一化 (Apply Norm)
    scale_factors = {
        "Agent-agent collision rate": 1.0,
        "Agent-lanelet collision rate": 1.0,
        "Total collision rate": 1.5,
        "Relative centerline deviation": 1.0,
        "Relative average speed": 1.5,
        "Smoothness": 1.5,
    }

    def calculate_optimized_norm(row):
        metric_name = row["Metric"]
        metric_lower = metric_name.lower()
        g_min = row["Global_Min"]
        g_max = row["Global_Max"]
        current_val = row["Value"]

        range_val = g_max - g_min
        if range_val <= 0:
            base_norm = 0.5
        else:
            scale = scale_factors.get(metric_name, 1.0)
            expanded_range = range_val * scale
            mid = 0.5 * (g_min + g_max)
            L = mid - 0.5 * expanded_range
            U = mid + 0.5 * expanded_range
            denom = U - L
            if denom <= 0:
                base_norm = 0.5
            else:
                base_norm = (current_val - L) / denom
                if base_norm < 0.0:
                    base_norm = 0.0
                elif base_norm > 1.0:
                    base_norm = 1.0

        is_negative = any(keyword in metric_lower for keyword in negative_keywords)

        if is_negative:
            return 1.0 - base_norm
        else:
            return base_norm

    merged_df["Norm_Value"] = merged_df.apply(calculate_optimized_norm, axis=1)

    safety_scores = merged_df[merged_df["Metric"] == "Total collision rate"][
        ["Scenario", "Algorithm", "Norm_Value"]
    ].rename(columns={"Norm_Value": "Safety_Score"})

    merged_df = pd.merge(
        merged_df,
        safety_scores,
        on=["Scenario", "Algorithm"],
        how="left",
    )

    merged_df["Collision_Severity"] = 1.0 - merged_df["Safety_Score"]
    merged_df["Collision_Severity"] = merged_df["Collision_Severity"].fillna(0.0)
    merged_df["Collision_Severity"] = merged_df["Collision_Severity"].clip(0.0, 1.0)

    penalty_efficiency = 0.0
    penalty_smoothness = 0.0

    def apply_collision_penalty(row):
        norm_val = row["Norm_Value"]
        sev = row["Collision_Severity"]
        metric = row["Metric"]
        if metric == "Relative average speed":
            val = norm_val - penalty_efficiency * sev
            if val < 0.0:
                val = 0.0
            return val
        if metric == "Smoothness":
            val = norm_val - penalty_smoothness * sev
            if val < 0.0:
                val = 0.0
            return val
        return norm_val

    merged_df["Norm_Value"] = merged_df.apply(apply_collision_penalty, axis=1)

    safety_norm = merged_df[merged_df["Metric"] == "Total collision rate"][
        ["Scenario", "Algorithm", "Norm_Value"]
    ].rename(columns={"Norm_Value": "S"})

    efficiency_norm = merged_df[merged_df["Metric"] == "Relative average speed"][
        ["Scenario", "Algorithm", "Norm_Value"]
    ].rename(columns={"Norm_Value": "E"})

    se_merge = pd.merge(
        safety_norm,
        efficiency_norm,
        on=["Scenario", "Algorithm"],
        how="inner",
    )

    alpha = 0.7
    se_merge["Combined_Score"] = se_merge["S"].clip(0.0, 1.0) ** alpha * se_merge[
        "E"
    ].clip(0.0, 1.0) ** (1.0 - alpha)

    combined_rows = se_merge[["Scenario", "Algorithm", "Combined_Score"]].copy()
    combined_rows["Metric"] = "Safety-efficiency score"
    combined_rows["Stat"] = "Norm"
    combined_rows.rename(columns={"Combined_Score": "Value"}, inplace=True)

    # 4. 构建 Norm 行
    norm_data = merged_df[["Scenario", "Metric", "Algorithm"]].copy()
    norm_data["Stat"] = "Norm"
    norm_data["Value"] = merged_df["Norm_Value"]

    # 5. 合并回总表
    df_final = pd.concat([df, norm_data, combined_rows], ignore_index=True)

    # ---------------------------------------------------------
    # 格式设置 (Format)
    # ---------------------------------------------------------
    scenario_order = ["Clover", "Weave", "Merge", "Bypass"]
    df_final["Scenario"] = pd.Categorical(
        df_final["Scenario"], categories=scenario_order, ordered=True
    )

    metric_order = [
        "Agent-agent collision rate",
        "Agent-lanelet collision rate",
        "Total collision rate",
        "Relative centerline deviation",
        "Relative average speed",
        "Smoothness",
        "Safety-efficiency score",
    ]
    df_final["Metric"] = pd.Categorical(
        df_final["Metric"], categories=metric_order, ordered=True
    )

    pivot = df_final.pivot_table(
        index=["Scenario", "Metric"], columns=["Algorithm", "Stat"], values="Value"
    )

    # 列重排
    algo_order = ["SigmaRL", "TSC-Net", "X-MARL"]
    stat_order = ["Mean", "Min", "Max", "Norm"]

    new_column_order = []
    new_column_names = []

    for algo in algo_order:
        for stat in stat_order:
            if (algo, stat) in pivot.columns:
                new_column_order.append((algo, stat))
                new_column_names.append(f"{algo} ({stat})")

    pivot = pivot[new_column_order]
    pivot.columns = new_column_names

    try:
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            pivot.to_excel(writer, sheet_name="Comparison")

            # 美化
            worksheet = writer.sheets["Comparison"]
            # 设置列宽
            for i, col in enumerate(pivot.reset_index().columns):
                col_letter = (
                    chr(65 + i) if i < 26 else chr(65 + i // 26 - 1) + chr(65 + i % 26)
                )
                worksheet.column_dimensions[col_letter].width = 16

            # 高亮 Norm 列 (可选，通过条件格式让数据更直观)
            from openpyxl.styles import PatternFill

            yellow_fill = PatternFill(
                start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
            )

            # 找到包含 "Norm" 字样的列并标黄
            for col_idx in range(1, worksheet.max_column + 1):
                header = worksheet.cell(row=1, column=col_idx).value
                if header and "(Norm)" in str(header):
                    for row_idx in range(1, worksheet.max_row + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill

        print(f"✅ Excel 表格已生成: {os.path.abspath(output_file)}")
        print("   逻辑应用:")
        print(f"   - 负向指标 (取互补值 1-x): {negative_keywords}")
        print("   - 正向指标 (保持 x): [Relative average speed]")
        print("   - 全局范围: Min(所有算法Min) ~ Max(所有算法Max)")

    except Exception as e:
        print(f"Error: {e}")

    return df_final


def plot_radar_from_norm(df_final, output_dir=None):
    if output_dir is None:
        output_dir = os.path.join(OUTPUT_DIR, "radar_plots")
    os.makedirs(output_dir, exist_ok=True)

    metrics = [
        "Total collision rate",
        "Relative average speed",
        "Smoothness",
    ]
    labels = ["Collision \n rate", "Average speed", "Smoothness"]

    norm_df = df_final[df_final["Stat"] == "Norm"].copy()
    scenarios = norm_df["Scenario"].unique()
    algorithms = norm_df["Algorithm"].unique()

    print(algorithms)

    angles = np.linspace(0, 2 * np.pi, len(metrics), endpoint=False)
    angles = np.concatenate([angles, angles[:1]])

    colors = {
        "SigmaRL": "#4C72B0",  # 柔和蓝
        "TSC": "#55A868",  # 柔和绿
        "XP-MARL": "#C44E52",  # 柔和红
        "MFPO": "#8172B2",  # 紫
    }

    for scenario in scenarios:
        df_s = norm_df[norm_df["Scenario"] == scenario]

        fig, ax = plt.subplots(subplot_kw={"polar": True}, figsize=(3.2, 3.5))
        ax.set_facecolor("#f5f6fa")
        fig.patch.set_facecolor("white")
        ax.grid(color="white", linewidth=1.2, alpha=1.0)
        ax.spines["polar"].set_visible(False)

        for alg in algorithms:
            df_sa = df_s[df_s["Algorithm"] == alg]
            if df_sa.empty:
                continue

            values = []
            for m in metrics:
                row = df_sa[df_sa["Metric"] == m]
                if row.empty:
                    values.append(0.0)
                else:
                    values.append(float(row["Value"].iloc[0]))

            values = values + values[:1]
            color = colors.get(alg, None)

            ax.plot(
                angles,
                values,
                label=alg,
                linewidth=2.0,
                marker="o",
                markersize=4,
                color=color,
            )
            ax.fill(angles, values, alpha=0.15, color=color)

        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(labels, fontsize=10)
        ax.set_yticks([0.3, 0.6, 0.9])
        ax.set_yticklabels(["0.3", "0.6", "0.9"])
        ax.set_ylim(0, 1)
        ax.set_title(str(scenario), fontsize=14, y=1.20)
        ax.tick_params(axis="x", pad=12)
        ax.legend(
            loc="upper right",
            bbox_to_anchor=(0.25, 0.70),
            fontsize=8,
            frameon=True,
            facecolor="white",
            framealpha=0.9,
        )

        fname = f"radar_{scenario}.png"
        path = os.path.join(output_dir, fname)
        plt.tight_layout()
        plt.savefig(path, dpi=200)
        plt.close(fig)


df_final = generate_excel_final_norm(log_content)
plot_radar_from_norm(df_final)
